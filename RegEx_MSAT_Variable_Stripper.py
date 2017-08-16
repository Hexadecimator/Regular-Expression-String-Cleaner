'''
Created on Jun 16, 2017

@author: logans

Patterns:

1. Variables start with "PFCC_" and end with "= 1"
2. NOT statements start with "NOT(" and end at the proceeding ")"

TODO:
1. do a first pass to remove the NOT statements:
   1a. Use RegEx to find anything starting with "NOT(" and the proceeding ")" and replace it with "<!!!NOT_STATEMENT_REMOVED!!!>" (initially - for debugging)
   1b. Store the updated equation logic without NOT statements inside the "NOTs_Removed" (idx=2) worksheet
2. Once all NOT statements are stripped out, strip each remaining variable into a list of individual variables:
   2a. Use RegEx to find anything starting with "PFCC_" and ending with "= 1" and give it its own entry on the list
   2b. When the list is fully populated, print each variable to its own row in the "Variable_List" (idx=3)

how to use RegEx to replace the matched part of a string:

    import re 
    re.sub('NOT(.*?)', '', string)    # Where "string" contains the current fault equation logic-holding cell being analyzed

'''
import re
import openpyxl as xl

def main():
    wb_original = xl.load_workbook('Global_MSAT_Signal_Analysis.xlsx')
    sheetnames = wb_original.get_sheet_names()
    
    # The next 2 "for" loops print debug information to confirm the correct excel file was opened
    count = 0
    for sht_idx in sheetnames:
        print("Sheet Index: " + str(count) + " Sheet Name: " + sht_idx)
        count += 1
        
    for sht_idx in sheetnames:
        print("# of rows inside " + sht_idx + " workbook: " + str(wb_original.get_sheet_by_name(sht_idx).max_row))    
    
    # hard-code the Test_Sheet workbook to be our working (content-holding) workbook
    original_worksheet = wb_original.get_sheet_by_name("Full_Sheet")
    NOTs_Removed_worksheet = wb_original.get_sheet_by_name("NOTs_Removed")
    variable_worksheet = wb_original.get_sheet_by_name("Variable_List")
    
    ow_num_rows = original_worksheet.max_row #ow = "original worksheet" abbreviated
    print(str(ow_num_rows))

    count = 1
    for rowidx in range(1,ow_num_rows+1):
        # this loads the cell coordinate variables for locating manipulating and storing the current cell's contents
        curr_Cell_Name_Coord = "A" + str(rowidx)
        curr_Cell_Content_Coord = "B" + str(rowidx)

        # this uses the cell coordinate variables to load the current cell's data for manipulation
        curr_Cell_Name = str(original_worksheet[curr_Cell_Name_Coord].value)
        curr_Cell_Content = str(original_worksheet[curr_Cell_Content_Coord].value)
        
        # re.sub will remove all instances of "NOT(*)" in the current equation's logic(where * is anything)
        NOT_Removed_Cell_Content = re.sub('ABSENT(.*?)\)', '', curr_Cell_Content) #!*!*!NEW!*!*!
        NOT_Removed_Cell_Content = re.sub('NOT(.*?)\)', '', NOT_Removed_Cell_Content)
        
        # the 2 print statements below are for debug
        #print("Equation # " + str(count) + " named " + curr_Cell_Name + " contains the following modified content:")
        #print(NOT_Removed_Cell_Content + "\n")
        
        # load the new equation logic with all the NOT(*) and ABSENT(*) statements removed into a new worksheet
        NOTs_Removed_worksheet[curr_Cell_Name_Coord] = curr_Cell_Name
        NOTs_Removed_worksheet[curr_Cell_Content_Coord] = NOT_Removed_Cell_Content
        
        count += 1
        
    nr_num_rows = NOTs_Removed_worksheet.max_row # nr = NOTs Removed abbreviated
    
    for rowidx in range(1,nr_num_rows+1):
        curr_Cell_Name_Coord = "A" + str(rowidx)
        curr_Cell_Content_Coord = "B" + str(rowidx)

        curr_Cell_Name = str(NOTs_Removed_worksheet[curr_Cell_Name_Coord].value)
        curr_Cell_Content = str(NOTs_Removed_worksheet[curr_Cell_Content_Coord].value)
        
        m = re.search('PFCC_(.+)= 1', curr_Cell_Content)
        if m:
            found = m.group(0)
            
            # I don't know how to do this better, so brute force clean up!
            while ')' in found:
                found = found.replace(')','')
                
            while '(' in found:
                found = found.replace('(','')
                
            while 'AND' in found:
                found = found.replace('AND','')
                
            while 'OR' in found:
                found = found.replace('OR','')
                
            while 'Pri Flight Control Computer 1 1.' in found:
                found = found.replace('Pri Flight Control Computer 1 1.','')
            
            while 'Pri Flight Control Computer 2 2.' in found:
                found = found.replace('Pri Flight Control Computer 2 2.','')
                
            while 'Pri Flight Control Computer 3 3.' in found:
                found = found.replace('Pri Flight Control Computer 3 3.','')
                
            while '= 1' in found:
                found = found.replace('= 1','')
            
            while 'PFCC 1 1.' in found:
                found = found.replace('PFCC 1 1.','')
            
            while 'PFCC 2 2.' in found:
                found = found.replace('PFCC 2 2.','')
            
            while 'PFCC 3 3.' in found:
                found = found.replace('PFCC 3 3.','')
                
            while 'NULL' in found:
                found = found.replace('NULL','')
                
            while 'FBW IIM MAINT FAULTS 6' in found:
                found = found.replace('FBW IIM MAINT FAULTS 6','')
                
            while 'FBW PFCC MAINT FAULTS 9' in found:
                found = found.replace('FBW PFCC MAINT FAULTS 9','')
            
            while 'FBW IIM MAINT FAULTS 2' in found:
                found = found.replace('FBW IIM MAINT FAULTS 2','')
                
            while 'FBW IIM MAINT FAULTS 3' in found:
                found = found.replace('FBW IIM MAINT FAULTS 3','')
                
            while 'FBW IIM MAINT FAULTS 5' in found:
                found = found.replace('FBW IIM MAINT FAULTS 5','')
            
            while 'FBW IIM MAINT FAULTS 4' in found:
                found = found.replace('FBW IIM MAINT FAULTS 4','')
                
            while 'FBW MISC REU MAINT FAULTS 3' in found:
                found = found.replace('FBW MISC REU MAINT FAULTS 3','')
            
            while 'Fault 1' in found:
                found = found.replace('Fault 1','Fault')
                
            while '2OF' in found:
                found = found.replace('2OF','')
            
            while '3OF' in found:
                found = found.replace('3OF','')
            
            while '5OF' in found:
                found = found.replace('5OF','')         
            
            while '1.' in found:
                found = found.replace('1.','')
             
            while '2.' in found:
                found = found.replace('2.','')
             
            while '3.' in found:
                found = found.replace('3.','')
                
            while ',' in found:
                found = found.replace(',','')
                
            while '  ' in found:
                found = found.replace('  ',' ')
                
            if "PFCC_" in found:
                found = found.replace('PFCC_', '\r\nPFCC_')
                
            if "PFCC-" in found:
                found = found.replace('PFCC-', '\r\nPFCC_')
            
        else:
            # there were no PFCC A429 variables found in the equation (means it was an R-C equation)
            print("Variables not found in NOTs_Removed worksheet @ cell: " + curr_Cell_Content_Coord)
        
        variable_worksheet[curr_Cell_Name_Coord] = curr_Cell_Name
        variable_worksheet[curr_Cell_Content_Coord] = found.strip()   
   
 
    # Save the data and close the excel file
    wb_original.save("Global_MSAT_Signal_Analysis.xlsx")
    print("Workbook Saved and Closed")
    
main()
